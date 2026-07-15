#!/usr/bin/env python3
"""
glad_load.py — normalise every store flow to a canonical FEDEFL identity, using the
official GLAD/FEDEFL nomenclature resources (DATABASE_PLAN.md §C, "proper" bridge).

Why: matching CFs to inventory flows by heuristic name/CAS keys only brackets the
hard categories (ecotoxicity ±13-15%) because the same substance appears under many
names/UUIDs across nomenclatures. FEDEFL is the authoritative "one list": it assigns
every substance a canonical *flowable* (with CAS + synonyms), and the GLAD mapped
files give the authoritative source-name→flowable correspondence PLUS the curated
NO_FLOW_MATCH exceptions (e.g. biogenic CO2) that heuristics get wrong.

fed_id  =  "<canonical FEDEFL flowable> | <compartment>"

Resolution per flow (first hit wins), context from the flow's own category:
  1. GLAD mapped files:  (source name, compartment) -> FEDEFL flowable   [most authoritative]
  2. FEDEFL master:      CAS -> flowable
  3. FEDEFL master:      name / synonym -> flowable
NO_FLOW_MATCH entries (by name+medium) set fed_id = NULL (never bridge — e.g. CO2 biogenic).

Run (one-off, after loading databases):
    python3 glad_load.py                # needs the two mapped files fetched (glad_fetch.py)
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl

from canonical_store import CanonicalStore, DEFAULT_DB
from flowkey import norm_name, norm_cas, fine_compartment, medium_of

GLAD = (Path(__file__).resolve().parent.parent
        / "data" / "References_C" / "GLAD-ElementaryFlowResources-master" / "Mapping")
MAPPED = GLAD / "Output" / "Mapped_files"
FEDEFL_CSV = GLAD / "Input" / "Flowlists" / "FEDEFLv1.0.3.csv"
MAPPED_FILES = ["ecoinventEFv3.7-FEDEFLv1.0.3.xlsx", "ILCD-EFv3.0-FEDEFLv1.0.3.xlsx"]

NOMAP_TYPES = {"NO_FLOW_MATCH_MANUAL", "NO_MAPPING"}


def _load_mapped(fn: str):
    """Return (src2flow, nomap): (norm_name, fine_ctx)->flowable ; set of (norm_name, medium)."""
    src2flow, src2flow_med, nomap = {}, {}, set()
    wb = openpyxl.load_workbook(str(MAPPED / fn), read_only=True, data_only=True)
    ws = wb.active
    hdr = [str(x) for x in next(ws.iter_rows(values_only=True))]
    i_sn, i_sc = hdr.index("SourceFlowName"), hdr.index("SourceFlowContext")
    i_mt = hdr.index("MapType")
    i_tn, i_tc = hdr.index("TargetFlowName"), hdr.index("TargetFlowContext")
    for r in ws.iter_rows(min_row=2, values_only=True):
        sn = norm_name(r[i_sn])
        if not sn:
            continue
        mt = str(r[i_mt] or "").strip()
        s_fine = fine_compartment(r[i_sc])
        if mt in NOMAP_TYPES:
            nomap.add((sn, medium_of(f"x|{s_fine}").split("|", 1)[1]))
            continue
        tn = norm_name(r[i_tn])
        if not tn:
            continue
        src2flow.setdefault((sn, s_fine), tn)
        src2flow_med.setdefault((sn, medium_of(f"x|{s_fine}").split("|", 1)[1]), tn)
    wb.close()
    return src2flow, src2flow_med, nomap


def _load_fedefl():
    """cas->flowable and name/synonym->flowable from the FEDEFL master list."""
    cas2flow, name2flow = {}, {}
    with open(FEDEFL_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            flow = norm_name(r["Flowable"])
            if not flow:
                continue
            c = norm_cas(r.get("CAS No"))
            if c:
                cas2flow.setdefault(c, flow)
            name2flow.setdefault(flow, flow)
            for syn in (r.get("Synonyms") or "").split(";"):
                s = norm_name(syn)
                if s:
                    name2flow.setdefault(s, flow)
    return cas2flow, name2flow


def build_fed_ids(store: CanonicalStore) -> dict:
    if not (MAPPED / MAPPED_FILES[0]).exists() or (MAPPED / MAPPED_FILES[0]).stat().st_size < 2000:
        print("ERROR: GLAD mapped files not present/real. Run: python3 glad_fetch.py", file=sys.stderr)
        raise SystemExit(1)

    print("loading GLAD mapped files ...", flush=True)
    src2flow, src2flow_med, nomap = {}, {}, set()
    for fn in MAPPED_FILES:
        a, b, c = _load_mapped(fn)
        src2flow.update(a); src2flow_med.update(b); nomap |= c
    print(f"  source→flowable (fine): {len(src2flow)}  medium: {len(src2flow_med)}  NO_FLOW_MATCH: {len(nomap)}")
    print("loading FEDEFL master list ...", flush=True)
    cas2flow, name2flow = _load_fedefl()
    print(f"  CAS→flowable: {len(cas2flow)}  name/synonym→flowable: {len(name2flow)}")

    # migrate: add fed_id column
    cols = {r[1] for r in store.conn.execute("PRAGMA table_info(flows)")}
    if "fed_id" not in cols:
        store.conn.execute("ALTER TABLE flows ADD COLUMN fed_id TEXT")
    store.conn.execute("CREATE INDEX IF NOT EXISTS ix_flows_fedid ON flows(fed_id)")

    stats = {"mapped": 0, "cas": 0, "name": 0, "nomap": 0, "unresolved": 0}
    updates = []
    for r in store.conn.execute(
        "SELECT uid, name, cas, category FROM flows WHERE flow_type='ELEMENTARY_FLOW'"
    ):
        nm = norm_name(r["name"]); fine = fine_compartment(r["category"])
        med = medium_of(f"x|{fine}").split("|", 1)[1]
        if (nm, med) in nomap:
            # sentinel: authoritatively "do not map" (e.g. biogenic CO2). Distinct from
            # NULL/unresolved so the solver blocks bridging instead of falling back.
            updates.append(("__NOMAP__", r["uid"])); stats["nomap"] += 1
            continue
        flow = src2flow.get((nm, fine)) or src2flow_med.get((nm, med))
        if flow:
            stats["mapped"] += 1
        else:
            c = norm_cas(r["cas"])
            flow = cas2flow.get(c) if c else None
            if flow:
                stats["cas"] += 1
            else:
                flow = name2flow.get(nm)
                if flow:
                    stats["name"] += 1
        if not flow:
            stats["unresolved"] += 1
            continue
        updates.append((f"{flow}|{fine}", r["uid"]))

    store.conn.executemany("UPDATE flows SET fed_id=? WHERE uid=?", updates)
    store.conn.commit()
    resolved = stats["mapped"] + stats["cas"] + stats["name"]
    total = resolved + stats["unresolved"] + stats["nomap"]
    print(f"\nfed_id assigned to {resolved}/{total} elementary flows "
          f"({resolved/max(total,1):.0%})  "
          f"[mapped {stats['mapped']}, CAS {stats['cas']}, name {stats['name']}]")
    print(f"  NO_FLOW_MATCH (excluded): {stats['nomap']}   unresolved: {stats['unresolved']}")
    return stats


if __name__ == "__main__":
    with CanonicalStore(DEFAULT_DB) as store:
        build_fed_ids(store)
